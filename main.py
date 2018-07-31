import smtplib
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import textile

file = 'test.xlsx' #xl flie
wb = openpyxl.load_workbook(file)
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])


email_cell = int(2)
subject_cell = int(3)
get_body = '1' #enter  0 if body is in xl or enter 1 if body is in txt
body_cell = int(4)
status_cell = sheet.max_row +1 
html_temp = True
file_body = open('body.txt').read()
signature = open('signature.html').read()
try:
    file_body = open('body.txt').read()
except:pass




class mailing():
    def __init__(self,email_id, password):
        self.email_id = str(email_id)
        self.password = str(password)
        
        
        

    def login(self):
        self.mail = smtplib.SMTP('smtp.gmail.com', 587)
        self.mail.ehlo()
        self.mail.starttls()
        self.mail.login(self.email_id, self.password)

    def make_mail(self,from_,to,subject,body):
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = from_
        msg['To'] = to
        body_ =  MIMEText(body, 'plain')
        msg.attach(body_)
        if html_temp == True :
            print('parsing html')
            html = MIMEText(body, 'html')
            msg.attach(html)
        self.mail.sendmail(from_, to, msg.as_string())



e_mail = mailing('ayush.tetrahedron@gmail.com','password')
e_mail.login()

max_rows = sheet.max_row
for i in range(2,max_rows+1):
        email = str( sheet.cell(row=i, column=2).value)
        subject = str(sheet.cell(row=i, column=3).value)
        if str(get_body) == '1':
            body = file_body
        else:
            body = str(sheet.cell(row=i, column=4).value)
        status = str(sheet.cell(row=i, column=5).value)
        body = textile.textile( body )
        body = body+'</br>'+signature
        #print(email,subject,body,status)
        if status != '2' :
            #send mail
            e_mail.make_mail(from_='Ayush',to=email,subject=subject,body=body)
            s = sheet.cell(row=i, column=5)
            s.value= '1'
            print('sent ')
            wb.save(file)
        else:
            pass

        
        
